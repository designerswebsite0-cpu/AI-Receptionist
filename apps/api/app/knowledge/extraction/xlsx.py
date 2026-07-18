from io import BytesIO

from openpyxl import load_workbook

from app.errors import ValidationErrorApp
from app.knowledge.extraction.base import ExtractedContent

# Spreadsheet-dimension limits (docs/phase-3/IMPLEMENTATION_PLAN.md's file
# validation rules) — guards against adversarial or accidental
# multi-million-row files consuming unbounded memory during extraction.
_MAX_ROWS_PER_SHEET = 50_000
_MAX_SHEETS = 50


def extract_xlsx(content: bytes) -> ExtractedContent:
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    if len(workbook.worksheets) > _MAX_SHEETS:
        raise ValidationErrorApp(f"Workbook has {len(workbook.worksheets)} sheets, exceeding the {_MAX_SHEETS} limit")

    tables = []
    text_parts = []
    for sheet in workbook.worksheets:
        rows: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            if all(cell is None for cell in row):
                continue
            if len(rows) >= _MAX_ROWS_PER_SHEET:
                raise ValidationErrorApp(
                    f"Sheet '{sheet.title}' has more than {_MAX_ROWS_PER_SHEET} rows"
                )
            row_values = ["" if cell is None else str(cell) for cell in row]
            rows.append(row_values)
            text_parts.append(" | ".join(row_values))
        tables.append({"sheet_name": sheet.title, "rows": rows})

    raw_text = "\n".join(text_parts)
    return ExtractedContent(
        raw_text=raw_text,
        extraction_method="openpyxl",
        word_count=len(raw_text.split()),
        tables=tables,
    )
